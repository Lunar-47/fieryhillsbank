from flask import Flask, render_template, request, redirect, url_for, session, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from datetime import date
import pandas as pd
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from openpyxl import Workbook
import re

app = Flask(__name__)
app.secret_key = "sdhfaushdfkjlhsdkfb"

# -----------------------------
# Database functions
# -----------------------------
def get_db_connection():
    conn = sqlite3.connect("hawkeyes.db")
    conn.row_factory = sqlite3.Row
    return conn

def get_students():
    conn = get_db_connection()
    students = conn.execute("SELECT * FROM students").fetchall()
    conn.close()
    return students

def get_transactions(student_id):
    conn = get_db_connection()
    transactions = conn.execute(
        "SELECT * FROM transactions WHERE student_id = ? ORDER BY date",
        (student_id,)
    ).fetchall()
    conn.close()
    return transactions

def calculate_balance(transactions):
    balance = sum(t["credit"] for t in transactions) - sum(t["debit"] for t in transactions)
    return balance

# -----------------------------
# Routes
# -----------------------------
@app.route("/", methods=["GET"])
def home():
    if isinstance((session.get("user_type")), str) or session.get("user_type") < 1:
        return redirect(url_for("login"))

    search_query = request.args.get("q")
    conn = get_db_connection()
    
    if search_query:
        students = conn.execute(
            "SELECT * FROM students name LIKE ?", ('%' + search_query + '%',)
        ).fetchall()
    else:
        students = conn.execute("SELECT * FROM students").fetchall()
    conn.close()
    return render_template("home.html", students=students, search_query=search_query)

@app.route("/balance/<int:student_id>")
def balance(student_id):
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))
    
    conn = get_db_connection()
    transactions = conn.execute(
        "SELECT * FROM transactions WHERE student_id = ? ORDER BY date DESC",  # sort newest first
        (student_id,)
    ).fetchall()
    
    student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    conn.close()

    # Calculate the balance (sum of credits - debits)
    balance = sum(t["credit"] - t["debit"] for t in transactions)

    # Render the balance template with the correct context
    return render_template(
        "balance.html",
        student_name=student["name"],  # passing student name as 'student_name'
        student_id=student_id,  # passing student_id
        transactions=transactions,  # passing the list of transactions
        balance=balance,  # passing the calculated balance
        grade=student["grade"]  # passing the student's grade level without "Grade_"
    )

@app.route("/add_transaction/<int:student_id>", methods=["GET", "POST"])
def add_transaction(student_id):
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))
    
    conn = get_db_connection()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    
    if not student:
        conn.close()
        return "<h2>Student not found.</h2>", 404

    if request.method == "POST":
        date_val = request.form["date"]
        description = request.form["description"]
        debit = int(request.form["debit"] or 0)
        credit = int(request.form["credit"] or 0)

        conn.execute(
            "INSERT INTO transactions (student_id, date, description, debit, credit) VALUES (?, ?, ?, ?, ?)",
            (student_id, date_val, description, debit, credit)
        )
        conn.commit()
        conn.close()
        return redirect(url_for("balance", student_id=student_id))

    conn.close()
    return render_template("add_transaction.html", student=student)

@app.route("/delete_transaction/<int:transaction_id>", methods=["POST"])
def delete_transaction(transaction_id):
    conn = get_db_connection()
    tx = conn.execute("SELECT * FROM transactions WHERE id=?", (transaction_id,)).fetchone()
    if not tx:
        conn.close()
        return redirect(url_for("home"))
    
    student_id = tx["student_id"]
    conn.execute("DELETE FROM transactions WHERE id=?", (transaction_id,))
    conn.commit()
    conn.close()
    
    return redirect(url_for("balance", student_id=student_id))

@app.route("/edit_transaction/<int:transaction_id>", methods=["GET", "POST"])
def edit_transaction(transaction_id):
    conn = get_db_connection()
    # Fetch the transaction
    transaction = conn.execute(
        "SELECT * FROM transactions WHERE id = ?", (transaction_id,)
    ).fetchone()

    if not transaction:
        conn.close()
        return "<h2>Transaction not found</h2>", 404

    # Fetch the student for display
    student = conn.execute(
        "SELECT * FROM students WHERE id = ?", (transaction["student_id"],)
    ).fetchone()

    if not student:
        conn.close()
        return "<h2>Student not found</h2>", 404

    if request.method == "POST":
        date_val = request.form["date"]
        description = request.form["description"]
        debit = int(request.form.get("debit") or 0)
        credit = int(request.form.get("credit") or 0)

        conn.execute(
            "UPDATE transactions SET date=?, description=?, debit=?, credit=? WHERE id=?",
            (date_val, description, debit, credit, transaction_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for("balance", student_id=student["id"]))

    conn.close()
    return render_template(
        "edit_transaction.html",
        transaction=transaction,
        student=student
    )

# -----------------------------
# Student Management
# -----------------------------
@app.route("/add_student", methods=["GET", "POST"])
def add_student():
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    message = ""

    if request.method == "POST":
        name = request.form.get("name").strip()
        grade = request.form.get("grade")
        password = request.form.get("password").strip()

        if not name or not grade or not password:
            message = "Name, grade, and password are all required!"
        else:
            # Hash the password
            hashed_password = generate_password_hash(password)

            # Check if student already exists
            existing = conn.execute("SELECT * FROM students WHERE name = ?", (name,)).fetchone()
            if existing:
                message = f"Student '{name}' already exists!"
            else:
                # Insert new student
                conn.execute(
                    "INSERT INTO students (name, grade, password) VALUES (?, ?, ?)",
                    (name, grade, hashed_password)
                )
                conn.commit()
                conn.close()
                return redirect(url_for("home"))

    conn.close()
    return render_template("add_student.html", message=message)
            
@app.route("/add_admin", methods=["GET", "POST"])
def add_admin():
    if session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    message = ""

    if request.method == "POST":
        name = request.form.get("name").strip()
        password = request.form.get("password").strip()
        permission = request.form.get("permission")

        if not name or not password or not permission:
            message = "Name, password, and permission level are required!"
        else:
            # Hash the password
            hashed_password = generate_password_hash(password)

            # Check if admin already exists
            existing = conn.execute("SELECT * FROM admins WHERE name = ?", (name,)).fetchone()
            if existing:
                message = f"Admin '{name}' already exists!"
            else:
                # Insert new student
                conn.execute(
                    "INSERT INTO admins (name, password, permissions) VALUES (?, ?, ?)",
                    (name, hashed_password, permission)
                )
                conn.commit()
                conn.close()
                return redirect(url_for("home"))

    conn.close()
    return render_template("add_admin.html", message=message)

@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    if session.get("user_type") != "student":
        return redirect(url_for("login"))

    message = ""
    student_id = session.get("user_id")

    if request.method == "POST":
        old_password = request.form.get("old_password").strip()
        new_password = request.form.get("new_password").strip()
        confirm_new_password = request.form.get("confirm_new_password").strip()

        # Validate the passwords
        if not new_password or not confirm_new_password:
            message = "New password is required!"
        elif new_password != confirm_new_password:
            message = "New passwords do not match!"
        else:
            conn = get_db_connection()
            student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()

            # Check if the old password is correct
            if student and check_password_hash(student["password"], old_password):
                # Hash the new password and update in the database
                hashed_new_password = generate_password_hash(new_password)
                conn.execute(
                    "UPDATE students SET password = ? WHERE id = ?",
                    (hashed_new_password, student_id)
                )
                conn.commit()
                conn.close()
                message = "Password changed successfully!"
            else:
                message = "Incorrect old password!"

    return render_template("change_password.html", message=message)

@app.route("/reset_student_password/<int:student_id>", methods=["GET", "POST"])
def reset_student_password(student_id):
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    message = ""
    if request.method == "POST":
        new_password = request.form.get("new_password").strip()
        confirm_new_password = request.form.get("confirm_new_password").strip()

        if not new_password or not confirm_new_password:
            message = "New password is required!"
        elif new_password != confirm_new_password:
            message = "New passwords do not match!"
        else:
            conn = get_db_connection()
            student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()

            if student:
                hashed_new_password = generate_password_hash(new_password)
                conn.execute(
                    "UPDATE students SET password = ? WHERE id = ?",
                    (hashed_new_password, student_id)
                )
                conn.commit()
                conn.close()
                message = f"Password for {student['name']} has been reset successfully!"
            else:
                message = "Student not found!"

    return render_template("reset_student_password.html", message=message, student_id=student_id)

@app.route("/payroll", methods=["GET", "POST"])
def payroll():
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    students = conn.execute("SELECT * FROM students").fetchall()

    if request.method == "POST":
        date_val = request.form["date"]
        for student in students:
            amount = int(request.form[f"payroll_{student['id']}"])

            # Save this amount as their default for next time
            conn.execute("UPDATE students SET default_payroll = ? WHERE id = ?", (amount, student["id"]))

            # Log the transaction
            conn.execute(
                "INSERT INTO transactions (student_id, date, description, debit, credit) VALUES (?, ?, ?, ?, ?)",
                (student["id"], date_val, "Payroll", 0, amount)
            )

        conn.commit()
        conn.close()
        return redirect(url_for("home"))

    # Sort students by name (case-insensitive)
    students = sorted(students, key=lambda s: s['name'].lower())

    conn.close()
    return render_template("payroll.html", students=students)

@app.route("/admin_accounts")
def admin_accounts():
    if session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    admin_account = conn.execute(
        "SELECT * FROM admins",
    ).fetchall()
    conn.close()

    return render_template(
        "admin_accounts.html",
        admin_account=admin_account
    )

# -----------------------------
# Excel Export / Upload
# -----------------------------
@app.route("/export_excel_by_grade")
def export_excel_by_grade():
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    students = conn.execute("SELECT * FROM students").fetchall()
    transactions = conn.execute("SELECT * FROM transactions ORDER BY date").fetchall()
    conn.close()

    # Gather grades and order them (K first)
    all_grades = set(str(s["grade"]) for s in students)
    
    # Extract numeric part from 'Grade_X' and sort
    grades_ordered = sorted([g for g in all_grades])

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for grade in grades_ordered:
            grade_students = [s for s in students if str(s["grade"]) == grade]
            if not grade_students:
                continue

            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="xlsxwriter") as writer:
                workbook = writer.book

                # --- Formats ---
                header_format = workbook.add_format({
                    "bold": True,
                    "bottom": 2,  # thick bottom for headers
                    "align": "center",
                    "valign": "vcenter"
                })
                right_border_format = workbook.add_format({
                    "bold": True,
                    "bottom": 2,
                    "right": 2,
                    "align": "center",
                    "valign": "vcenter"
                })
                date_format = workbook.add_format({"num_format": "yyyy/mm/dd"})

                # Debit & Credit header formats
                debit_header_format = workbook.add_format({
                    "bold": True,
                    "bottom": 2,
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": "#ff7c80",
                    "font_color": "#000000"
                })
                credit_header_format = workbook.add_format({
                    "bold": True,
                    "bottom": 2,
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": "#a9d08e",
                    "font_color": "#000000"
                })

                # Debit & Credit cell formats with thin top/bottom borders
                debit_cell_format = workbook.add_format({
                    "bg_color": "#ff7c80",
                    "font_color": "#000000",
                    "top": 1,
                    "bottom": 1
                })
                credit_cell_format = workbook.add_format({
                    "bg_color": "#a9d08e",
                    "font_color": "#000000",
                    "top": 1,
                    "bottom": 1
                })

                for student in grade_students:
                    student_tx = [t for t in transactions if t["student_id"] == student["id"]]

                    # Build the date column
                    formatted_dates = []
                    for t in student_tx:
                        raw_date = t["date"]
                        if not raw_date:
                            formatted_dates.append("")
                            continue
                        parsed_date = None
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                            try:
                                parsed_date = datetime.strptime(raw_date, fmt)
                                break
                            except Exception:
                                continue
                        formatted_dates.append(parsed_date.strftime("%Y/%m/%d") if parsed_date else raw_date)

                    if student_tx:
                        df = pd.DataFrame({
                            "Date": formatted_dates,
                            "Description": [t["description"] for t in student_tx],
                            "Debit": [t["debit"] for t in student_tx],
                            "Credit": [t["credit"] for t in student_tx]
                        })
                    else:
                        df = pd.DataFrame(columns=["Date", "Description", "Debit", "Credit"])

                    # Pad to 500 rows
                    while len(df) < 500:
                        df.loc[len(df)] = ["", "", "", ""]

                    # Balance formulas
                    df["Balance"] = ""
                    for i in range(1, 501):
                        if i == 1:
                            df.at[i - 1, "Balance"] = f"=D{i+1}-C{i+1}"
                        else:
                            df.at[i - 1, "Balance"] = f"=E{i}+D{i+1}-C{i+1}"

                    sheet_name = student["name"][:31]
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]

                    # Write headers
                    headers = ["DATE", "DESCRIPTION", "DEBIT (-)", "CREDIT (+)", "BALANCE:"]
                    for col_idx, header_text in enumerate(headers):
                        if col_idx == 2:
                            worksheet.write(0, col_idx, header_text, debit_header_format)
                        elif col_idx == 3:
                            worksheet.write(0, col_idx, header_text, credit_header_format)
                        else:
                            worksheet.write(0, col_idx, header_text, header_format)

                    # Set column widths and formats
                    worksheet.set_column("A:A", 12, date_format)
                    worksheet.set_column("B:B", 30)  # Description
                    worksheet.set_column("C:C", 12, debit_cell_format)
                    worksheet.set_column("D:D", 12, credit_cell_format)
                    worksheet.set_column("E:E", 12)  # Balance

                    # Total formula
                    worksheet.write_formula("F1", "=E500", right_border_format)

            excel_buffer.seek(0)
            zip_file.writestr(f"Grade_{grade}.xlsx", excel_buffer.read())

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        download_name="Hawkeyes_By_Grade.zip",
        as_attachment=True
    )

@app.route("/upload_excel_zip", methods=["GET", "POST"])
def upload_excel_zip():
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    if request.method == "POST":
        if "file" not in request.files:
            return "<h2>No file part</h2>", 400
        file = request.files["file"]
        if file.filename == "":
            return "<h2>No selected file</h2>", 400

        try:
            zip_file = zipfile.ZipFile(file)
            conn = get_db_connection()

            for filename in zip_file.namelist():
                if not filename.lower().endswith(".xlsx"):
                    continue

                excel_data = zip_file.read(filename)
                wb = openpyxl.load_workbook(io.BytesIO(excel_data))

                # Extract grade from filename (e.g., "Grade_1", "Grade_5", etc.)
                grade_match = re.match(r"Grade_(\d+)", filename)
                if not grade_match:
                    continue  # Skip files that don't follow the "Grade_X" pattern

                grade = f"{grade_match.group(1)}"  # This will extract the grade (e.g., "Grade_1")

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # Match student by full name (sheet_name corresponds to student name)
                    student = conn.execute(
                        "SELECT * FROM students WHERE name=?",
                        (sheet_name,)
                    ).fetchone()

                    # If student doesn't exist, create a new student with default password
                    if not student:
                        # Create a new student record with the default password "Eaya2025"
                        default_password = "Eaya2025"
                        hashed_password = generate_password_hash(default_password)

                        conn.execute(
                            "INSERT INTO students (name, grade, password) VALUES (?, ?, ?)",
                            (sheet_name, grade, hashed_password)
                        )
                        conn.commit()

                        # Now fetch the newly created student
                        student = conn.execute(
                            "SELECT * FROM students WHERE name=?",
                            (sheet_name,)
                        ).fetchone()

                    # Iterate over rows in Excel sheet, preserving order
                    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                        date_val, description, debit, credit = row
                        if date_val is None:
                            continue

                        # Convert to YYYY-MM-DD
                        if isinstance(date_val, (datetime, date)):
                            date_val = date_val.strftime("%Y-%m-%d")
                        elif isinstance(date_val, str):
                            parsed_date = None
                            for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                                try:
                                    parsed_date = datetime.strptime(date_val, fmt)
                                    break
                                except ValueError:
                                    continue
                            if parsed_date:
                                date_val = parsed_date.strftime("%Y-%m-%d")
                            else:
                                continue
                        else:
                            continue

                        # Clean and convert debit/credit values
                        debit = str(debit or 0).strip().replace('\xa0', '').replace(',', '')
                        debit = int(debit) if debit and debit.isdigit() else 0

                        credit = str(credit or 0).strip().replace('\xa0', '').replace(',', '')
                        credit = int(credit) if credit and credit.isdigit() else 0

                        # Insert transaction for the student
                        conn.execute(
                            "INSERT INTO transactions (student_id, date, description, debit, credit) VALUES (?, ?, ?, ?, ?)",
                            (student["id"], date_val, description, debit, credit)
                        )

            conn.commit()
            conn.close()
            return redirect(url_for("home"))

        except Exception as e:
            return f"<h2>Error processing ZIP file:</h2><p>{e}</p>", 500

    return '''
    <h2>Upload ZIP File with Excel Sheets for All Grades</h2>
    <form method="post" enctype="multipart/form-data">
        <input type="file" name="file" accept=".zip" required>
        <button type="submit">Upload ZIP</button>
    </form>
    '''

# -----------------------------
# Student login/logout
# -----------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    message = ""

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_db_connection()
        student = conn.execute(
            "SELECT * FROM students WHERE name = ?", (name,)
        ).fetchone()

        admin = conn.execute(
            "SELECT * FROM admins WHERE name = ?", (name,)
        ).fetchone()
        conn.close()

        # Check student login
        if student and check_password_hash(student["password"], password):
            session["user_type"] = "student"
            session["user_id"] = student["id"]
            return redirect(url_for("student_dashboard"))
        elif admin and check_password_hash(admin["password"], password):
            if admin["permissions"] > 0:
                session["user_type"] = admin["permissions"]
                return redirect(url_for("home"))
            else:
                message = "This account is not active"
        else:
            message = "Invalid name or password!"

    return render_template("login.html", message=message)


@app.route("/logout")
def logout():
    session.pop("user_type", None)
    session.pop("user_id", None)
    return redirect(url_for("login"))

@app.route("/student/dashboard")
def student_dashboard():
    if session.get("user_type") != "student":
        return redirect(url_for("login"))

    student_id = session["user_id"]
    conn = get_db_connection()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    transactions = conn.execute(
        "SELECT * FROM transactions WHERE student_id = ? ORDER BY date", (student_id,)
    ).fetchall()
    conn.close()

    balance_amount = calculate_balance(transactions)

    return render_template(
        "student_view.html",
        student=student,
        transactions=transactions,
        balance=balance_amount
    )

@app.route("/delete_student/<int:student_id>", methods=["POST"])
def delete_student(student_id):
    if session.get("user_type") != 1 and session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    # Delete all transactions for this student
    conn.execute("DELETE FROM transactions WHERE student_id = ?", (student_id,))
    # Delete the student record
    conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("home"))

@app.route("/delete_admin/<int:admin_id>", methods=["POST"])
def delete_admin(admin_id):
    if session.get("user_type") != 2:
        return redirect(url_for("login"))

    conn = get_db_connection()
    # Delete the admin record
    conn.execute("DELETE FROM admins WHERE id = ?", (admin_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("home"))

# -----------------------------
# Run the app
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)
